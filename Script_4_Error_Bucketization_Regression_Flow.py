#!/usr/bin/python3
import os
import sys
import re
import argparse
import subprocess
import shutil
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

#Getting error signatures from error.log to a dictionary
def read_error_signatures(input_error_path):
	signatures = {}
	pattern = r'(\w+)\s*=\s*"([^"]*)"'
	with open(input_error_path, 'r') as file:
		log_contents = file.read()
	matches = re.findall(pattern, log_contents)
	for key, value in matches:
		signatures[key] = value
	return signatures

input_error_path = 'input_error.log'	
signatures = read_error_signatures(input_error_path)	

#Getting error signatures to respective variables
TEST_START_SIGNATURE 				= signatures.get('TEST_START_SIGNATURE')
COMPILATION_SUCCESSFUL_SIGNATURE 	= signatures.get('COMPILATION_SUCCESSFUL_SIGNATURE')
TEST_PASS 							= signatures.get('TEST_PASS')
TEST_FAIL 							= signatures.get('TEST_FAIL')
SIMULATION_FINISH_SIGNATURE 		= signatures.get('TEST_START_SIGNATURE')
simulation_log_file 				= signatures.get('SIMULATION_LOG_FILE_NAME')
compilation_log_file 				= signatures.get('COMPILATION_LOG_FILE_NAME')
STAR_ERROR_SIGNATURE				= r"\*Error"

# Initialize error messages in a directory
errors_directory = "temp_files"
os.makedirs(errors_directory, mode=0o755, exist_ok=True)
errors_file_path = os.path.join(errors_directory, 'errors.txt')
path_input_path = os.path.join(errors_directory, 'root_path.txt')
path_dir_path = os.path.join(errors_directory, 'dir_path.txt')

#README.txt file content
readme_content="""
Script Name: regression_flow.py

Description:
This script traverse through log files, and generate a detailed EXCEL report with first error in compilation error, boot error, hung case.
It generates a detailed report containing count of all simulation errors.
It generates a report containing total occurances of compilation error, boot error, hung case and test failed.

Inputs needed:
1.Input all the signatures in input_error.log
2.After running the script it asks for regression directory ,give the path or default is pwd.

Usage:
	1.To categorize files in current directory and generate a EXCEL report:
		./regtesting.py		or		python regtesting.py
Arguments:

--help
	Print help information.

Output:
1. The EXCEL report is generated in a structured directory based on the current date and time:
   - Directory: excel_reports/<Year>/<Month>/<Day>/
   - Filename: error_report_<HH:MM:SS>.xlsx
2.report contains:
	-Detailed report: Contains organized error messages and counts
	-regression status: Contains counts of compilation error cases, boot error cases, hung cases and test failed cases.
	-4 sheets with testname,first error in compilation error, boot report, hung case respectively.

To run the script:
	python regtesting_custom.py or ./regtesting_custom.py
	
Scripts and input files included:
	1.regression_flow.py : To generate excel report.
	2.input_errors.log   : To input error signatures.


"""

#To print readme content
def print_readme():
	print(readme_content)
	
#parsing arguments
def help_arg():
	parser = argparse.ArgumentParser(description='Features and usage of the script', add_help=False)
	parser.add_argument('--help', action='store_true', help='Print help information')
	args = parser.parse_args()
	
	if args.help:
		print_readme()
		return

# List of predefined error messages
errors_list = [
	"UVM_WARNING @",
	"UVM_ERROR @",
	"UVM_FATAL @",
	"ASSERTION_ERROR @",
	"SIMULATION_TIMEOUT @",
	"CONSTRAINT_ERROR @",
	"RUNTIME_ERROR @"]
	
# Writing the error messages to errors.txt
try:
	with open(errors_file_path, 'w') as errors_file:
		for error in errors_list:
			errors_file.write(f"{error}\n") 
except IOError as e:
	print(f"Error writing to file '{errors_file_path}': {e}")
	
#Getting the regression directory path
def get_user_input():
	print("*********************** Regression Flow Script started *********************** \n")
	path = input("--> Enter Directory path : ").strip()
	print("	\n \n")
	if not path or path in [".", "./"]:
		path = os.getcwd()
	try:
		with open(path_input_path, 'w') as path_input_file:
			path_input_file.write(path)
		return path
	except IOError as e:
		print(f"Error writing to file '{path_input_path}': {e}")
		sys.exit(1)
		
# Finding folder paths containing log files 
def find_folder_path(root_path):
	matching_folders = []
	pattern = re.compile(r'test', re.IGNORECASE)
	for root, dirs, files in os.walk(root_path):
		for name in dirs:
			if pattern.search(name):	
				dir_path = os.path.join(root, name) 
				matching_folders.append(dir_path)
	matching_folders.sort()
	return matching_folders
	
#Write the given data to a specified file.
def write_to_file(file_path, data):
	with open(file_path, 'w') as file:
		for item in data:
			file.write(item + '\n')

#Read lines from a file and return a list
def read_lines(file_path):
	try:
		with open(file_path, 'r') as file:
			return [line.strip() for line in file.readlines()]
	except FileNotFoundError:
		print(f"Error: File '{file_path}' not found.")
		return []

#processing log files to find error signatures
def process_log_files(log_files_path, error_list):
	error_count = {'Errors': {error: {'Total errors': 0, 'Files': set()} for error in error_list},'starErrors':{} }
	sim_comp_data = {'compilation_failed': [],'hung_case': [],'boot_error': [],'test_failed': []}
	compilation_errors = {'Errors': {error: {'Total errors': 0, 'Files': set()} for error in error_list},'starErrors':{}}
	#lists containing the specified error cases names
	boot_error_cases = []
	hung_cases = []
	compilation_failed_cases = []
	no_test_logs_found_cases = []
	sim_success_cases = []
	#sets containg the first error in specifies cases
	first_error_in_test = {}
	boot_error_cases_first_error = {}
	hung_cases_first_error = {}
	compilation_failed_cases_first_error = {}
	no_test_logs_found_cases_first_error = {}
	sim_success_cases_first_error = {} 
	
	#patterns created for finding in files
	pattern_p_f = re.compile(rf'{re.escape(TEST_PASS)}|{re.escape(TEST_FAIL)}',re.IGNORECASE)
	pattern_build_success = re.compile(COMPILATION_SUCCESSFUL_SIGNATURE,re.IGNORECASE)
	pattern_running = re.compile(TEST_START_SIGNATURE,re.IGNORECASE)
	pattern_finish = re.compile(SIMULATION_FINISH_SIGNATURE,re.IGNORECASE)
	pattern_star_error = re.compile(STAR_ERROR_SIGNATURE,re.IGNORECASE)
	pattern_errors = re.compile('|'.join(error_list), re.IGNORECASE)
	
	#Finding errors and returning the details of the errors
	for dirpath in log_files_path:
		simulation_file_path = os.path.join(dirpath, simulation_log_file)
		compilation_file_path = os.path.join(dirpath, compilation_log_file)
		found_build_successful = False
		found_running_test = False
		found_pass_fail = False
		found_finish = False
		found_star_error =False
		final_folder_name = os.path.basename(dirpath)
		first_error_signature = ""
		if os.path.isfile(simulation_file_path):
			with open(simulation_file_path, 'r') as file:
				for line in file:
					error_match = pattern_errors.search(line)
					if pattern_build_success.search(line):
						found_build_successful = True
					if pattern_running.search(line):
						found_running_test = True
					if pattern_p_f.search(line):
						found_pass_fail = True
					if pattern_finish.search(line):
						found_finish = True
					if pattern_star_error.search(line):
						found_star_error = True
						if line not in error_count['starErrors']:
							error_count['starErrors'][line] = []
						error_count['starErrors'][line].append(final_folder_name)
					if error_match :
						error = error_match.group()
						error_count['Errors'][error]['Total errors'] += 1
						error_count['Errors'][error]['Files'].add(final_folder_name)
						if final_folder_name not in first_error_in_test:
							first_error_in_test[final_folder_name] = line
							
		#storing simulation successful cases and the 1'st error 
		if found_pass_fail and found_running_test and os.path.isfile(compilation_file_path) and os.path.isfile(simulation_file_path):
			sim_success_cases.append(f"{final_folder_name}")
			sim_success_cases_first_error[final_folder_name] = first_error_in_test.get(final_folder_name, "No error log")

		#storing test failed cases
		if not os.path.isfile(simulation_file_path) and not os.path.isfile(compilation_file_path):
			no_test_logs_found_cases.append(f"{final_folder_name}")
			sim_comp_data['test_failed'].append(f"{final_folder_name}")
			
		#storing compilation errors and the first compilation error file
		if (not found_build_successful) :
			if os.path.isfile(compilation_file_path):
				with open(compilation_file_path, 'r') as file:
					for line in file:
						error_match = pattern_errors.search(line)
						if error_match:
							error = error_match.group()
							compilation_errors['Errors'][error]['Total errors'] += 1
							compilation_errors['Errors'][error]['Files'].add(final_folder_name)
							if final_folder_name not in compilation_failed_cases_first_error:
								compilation_failed_cases.append(final_folder_name)
								compilation_failed_cases_first_error[final_folder_name] = line
								sim_comp_data['compilation_failed'].append(f"{final_folder_name}")
								
		#Storing boot error cases and 1st error	file	
		if not found_running_test and os.path.isfile(simulation_file_path) and os.path.isfile(compilation_file_path):
			boot_error_cases.append(f"{final_folder_name}")
			sim_comp_data['boot_error'].append(f"{final_folder_name}")
			
		#Storing hung cases and 1st error file	
		if not found_pass_fail and found_running_test and os.path.isfile(simulation_file_path) and os.path.isfile(compilation_file_path):
			hung_cases.append(f"{final_folder_name}")
			sim_comp_data['hung_case'].append(f"{final_folder_name}")
			
	#storing the first boot error
	for item in boot_error_cases:
		if item in first_error_in_test:
			boot_error_cases_first_error[item] = first_error_in_test[item]
			
	#storing the first hung case
	for item in hung_cases:
		if item in first_error_in_test:
			hung_cases_first_error[item] = first_error_in_test[item]
	#storing compilation failed cases
	for item in compilation_failed_cases:
		if item in first_error_in_test:
			compilation_failed_cases_first_error[item] = first_error_in_test[item]
	
	#storing no test logs found cases
	for item in no_test_logs_found_cases:
		if item in first_error_in_test:
			no_test_logs_found_cases_first_error[item] = first_error_in_test[item]
	
	#returning all the required outputs for categorizing
	return error_count,sim_comp_data,first_error_in_test,boot_error_cases_first_error,hung_cases_first_error,no_test_logs_found_cases_first_error,sim_success_cases_first_error,compilation_failed_cases,compilation_failed_cases_first_error
				
#adjusting column width in excel
def adjust_column_width(ws):
	for col in ws.columns:
		max_length = 0
		column = col[0].column_letter
		for cell in col:
			try:
				if len(str(cell.value)) > max_length:
					max_length = len(cell.value)
			except:
				pass
		adjusted_width = (max_length + 2)
		ws.column_dimensions[column].width = adjusted_width
		
#wrapping text in excel
def wrap_text(ws):
	for row in ws.iter_rows():
		for cell in row:
			cell.alignment = Alignment(wrap_text=True)
			
#creation of excel report
def create_excel_report(error_count,sim_comp_data):
	#data of detailed report
	data = []
	for error, info in error_count['Errors'].items():
		files_str = ', '.join(info['Files'])
		data.append({
			'Error type': error.split('@')[0].strip(),
			'Total error count': info['Total errors'],
			'Files': files_str
		})
	for line ,info in error_count['starErrors'].items():
		data.append({
			'Error type': line,
			'Total error count': len(info),
			'Files': ', '.join(info)
		})

	#data of regression status
	data2 = []
	for key, values in sim_comp_data.items():
		data2.append({
			'Issue type' : key,
			'Total occurances' : len(values)
			})
	#data of simulation error	
	data3 = []
	for folder ,info in sim_success_cases_first_error.items():
		data3.append({
			'Test name': folder,
			'First error': info
		})
	#data of boot error
	data4 = []
	for folder ,info in boot_error_cases_first_error.items():
		data4.append({
			'Test name': folder,
			'First error': info
		})	
	#data of test hung
	data5 = []
	for folder ,info in hung_cases_first_error.items():
		data5.append({
			'Test name': folder,
			'First error': info
		})		
	#data of compilation error
	data6 = []
	for folder ,info in compilation_failed_cases_first_error.items():
		data6.append({
			'Test name': folder,
			'First error': info
		})			
		
	# Creation of DataFrame
	df = pd.DataFrame(data)
	df_regression_status = pd.DataFrame(data2)
	df_simulation_error  = pd.DataFrame(data3)
	df_boot_error		 = pd.DataFrame(data4)
	df_hung 			 = pd.DataFrame(data5)
	df_compilation_error = pd.DataFrame(data6)
	
	# Generation of date stamped folder path
	current_datetime = datetime.now().strftime("%X")
	year = datetime.now().strftime("%Y")
	month = datetime.now().strftime("%b")
	date = datetime.now().strftime("%d")
	excel_folder = os.path.join(os.getcwd(), "excel_reports", year, month, date)
	os.makedirs(excel_folder, exist_ok=True)
	
	#excel filename and path generation
	excel_filename = f'error_report_{current_datetime}.xlsx'
	excel_filepath = os.path.join(excel_folder, excel_filename)

	#Creating multiple sheets with titles
	try:
		# Saving DataFrame to Excel with multiple sheets
		with pd.ExcelWriter(excel_filepath) as writer:
			df.to_excel(writer, sheet_name='detailed report', index=False)
			df_regression_status.to_excel(writer, sheet_name='regression status', index=False)
			df_compilation_error.to_excel(writer, sheet_name='compilation error', index=False)
			df_simulation_error.to_excel(writer, sheet_name='simulation error', index=False)
			df_boot_error.to_excel(writer, sheet_name='boot error', index=False)
			df_hung.to_excel(writer, sheet_name='test hung', index=False)
			
		#adjusting column length and wrapping text in excel sheet
		workbook = load_workbook(excel_filepath)
		for sheet_name in workbook.sheetnames:
			sheet = workbook[sheet_name]
			adjust_column_width(sheet)
			wrap_text(sheet)
		workbook.save(excel_filepath)
		folder_containing_report_path = f"{excel_folder}"+"/"+f"{excel_filename}"
		print(folder_containing_report_path)
		print(f"==> Genereated output File : {print_hyperlink(folder_containing_report_path)}")
		print("	\n \n*****************  Regression Flow Completed  *****************\n \n")
	except Exception as e:
		print(f"Error creating Excel file: {e}")

#Finding the folder paths and storing in dir_path.txt
def finding_log_folders_paths():
	test_directories_found = False
	if os.path.exists(path):
		log_folders = find_folder_path(path)
		if log_folders:
			write_to_file(path_dir_path, log_folders)
			test_directories_found = True
		else:
			print("No test directories found in the given path.")
	else:
		print("Provided path does not exist.")
	return test_directories_found

#making the spread sheet a hyperlink 
def print_hyperlink(path):
	# Create a hyperlink using ANSI escape codes
	hyperlink = f"\033]8;;file://{path}\033\\{path}\033]8;;\033\\"
	return hyperlink

#Deleting the temporary files created
def delete_folder(folder_path):
	try:
		shutil.rmtree(folder_path)
	except Exception as e:
		print(f"Failed to delete temp_files folder. Error: {e}")

#main test	
if __name__ == "__main__":
	os.system('clear')
	help_arg()
	path = get_user_input()
	test_directories_found = finding_log_folders_paths()
	
	if test_directories_found :
		error_list = read_lines(errors_file_path)
		log_files_path = read_lines(path_dir_path)
		error_count,sim_comp_data,first_error_in_test,boot_error_cases_first_error,hung_cases_first_error,no_test_logs_found_cases_first_error,sim_success_cases_first_error,compilation_failed_cases,compilation_failed_cases_first_error = process_log_files(log_files_path, error_list)
		create_excel_report(error_count,sim_comp_data)
	else:
		print("Failed to find the log directories")
	delete_folder("temp_files")		
	
#Developer_Rahul
